"""Compile the complete list of NAICS-4 codes missing from
CTE_NAICS_CODES that appear as top-N NAICS for direct-CTE SOCs.
Includes BLS-published NAICS titles for sanity-check readability.
"""

from __future__ import annotations

import csv
import os
from collections import defaultdict
from pathlib import Path

import openpyxl
from neo4j import GraphDatabase

from ontology.crosswalks import COE_DEMAND_PATH
from ontology import oes as _oes
from employers.edd_scrape import CTE_NAICS_CODES

TOP_N = 5


def _load_naics_titles() -> dict[str, str]:
    """Pull NAICS_TITLE from the OEWS NAICS-4 file."""
    wb = openpyxl.load_workbook(_oes.OES_NAICS4_PATH, read_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    cols = {c: i for i, c in enumerate(header)}
    titles: dict[str, str] = {}
    for row in rows:
        if not row: continue
        naics_raw = row[cols["NAICS"]]
        title = row[cols["NAICS_TITLE"]]
        if naics_raw is None or title is None: continue
        s = str(naics_raw).strip()
        if len(s) == 6 and s.isdigit():
            titles.setdefault(s[:4], str(title))
    wb.close()
    # Also fold in NAICS-3 and NAICS-2 file titles for fallback rollups.
    for path, keylen in [(_oes.OES_NAICS3_PATH, 3), (_oes.OES_NAICS2_PATH, 2)]:
        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header = next(rows)
        cols = {c: i for i, c in enumerate(header)}
        for row in rows:
            if not row: continue
            naics_raw = row[cols["NAICS"]]
            title = row[cols["NAICS_TITLE"]]
            if naics_raw is None or title is None: continue
            s = str(naics_raw).strip()
            # Keep canonical form (will be useful as fallback labels for
            # codes whose NAICS-4 row was not published).
            titles.setdefault(s, str(title))
        wb.close()
    return titles


def _load_education() -> dict[str, str]:
    edu: dict[str, str] = {}
    with open(COE_DEMAND_PATH, newline="") as f:
        for row in csv.DictReader(f):
            if row["SOC"] and row.get("Typical Entry Level Education"):
                edu.setdefault(row["SOC"], row["Typical Entry Level Education"].strip())
    return edu


def _band(level: str) -> str:
    if level in {"Postsecondary nondegree award", "Associate's degree"}: return "Strong CTE"
    if level in {"High school diploma or equivalent", "Some college, no degree",
                 "No formal educational credential"}: return "Moderate CTE"
    return "Other"


def main(out_path: str) -> None:
    edu_by_soc = _load_education()
    naics_titles = _load_naics_titles()
    _oes._ensure_loaded()

    soc_to_naics: dict[str, list[tuple[str, float]]] = defaultdict(list)
    for naics4, rows in _oes._socs_by_naics4.items():
        for r in rows:
            pct = r.get("pct_total") or 0.0
            if pct > 0:
                soc_to_naics[r["soc"]].append((naics4, pct))
    for soc in soc_to_naics:
        soc_to_naics[soc].sort(key=lambda x: -x[1])

    driver = GraphDatabase.driver(
        os.environ["NEO4J_URI"],
        auth=(os.environ["NEO4J_USERNAME"], os.environ["NEO4J_PASSWORD"]),
    )
    with driver.session() as session:
        cte_socs = session.run(
            "MATCH (o:Occupation) RETURN o.soc_code AS soc, o.title AS title"
        ).data()
    driver.close()

    # Aggregate per missing NAICS.
    missing: dict[str, dict] = {}
    for r in cte_socs:
        soc, title = r["soc"], r["title"]
        edu = edu_by_soc.get(soc, "")
        band = _band(edu)
        if band not in ("Strong CTE", "Moderate CTE"):
            continue
        ranked = soc_to_naics.get(soc, [])[:TOP_N]
        for naics, pct in ranked:
            if naics in CTE_NAICS_CODES: continue
            entry = missing.setdefault(naics, {
                "title": naics_titles.get(naics, "(no BLS title)"),
                "socs": [],
                "total_pct": 0.0,
                "bands": set(),
            })
            entry["socs"].append({
                "soc": soc, "title": title, "pct": pct, "band": band,
            })
            entry["total_pct"] += pct
            entry["bands"].add(band)

    out: list[str] = []
    out.append("# Missing NAICS-4 Codes — Sanity-Check List\n\n")
    out.append(
        f"Complete list of NAICS-4 codes that appear in the top-{TOP_N} "
        f"NAICS by pct_total for at least one direct-CTE SOC but are "
        f"NOT in the project's curated `CTE_NAICS_CODES` search list. "
        f"Sorted by number of direct-CTE SOCs the NAICS tops, then by "
        f"aggregate pct_total weight.\n\n"
        f"NAICS titles are pulled from the BLS OEWS NAICS_TITLE field "
        f"(2023 vintage, in-repo at `backend/ontology/data/oes_2023/`).\n\n"
    )

    out.append(f"## Summary\n\n")
    out.append(f"- Distinct missing NAICS-4 codes: **{len(missing)}**\n")
    out.append(f"- Total (NAICS, SOC) pairs missed: "
               f"**{sum(len(v['socs']) for v in missing.values())}**\n")
    out.append(f"- NAICS missing for ≥3 different direct-CTE SOCs: "
               f"**{sum(1 for v in missing.values() if len(v['socs']) >= 3)}**\n")
    out.append(f"- NAICS missing for Strong CTE specifically: "
               f"**{sum(1 for v in missing.values() if 'Strong CTE' in v['bands'])}**\n\n")

    ranked = sorted(missing.items(),
                    key=lambda kv: (-len(kv[1]["socs"]), -kv[1]["total_pct"]))

    out.append("## Full missing list\n\n")
    out.append("| NAICS-4 | Title | # CTE SOCs it tops | Σ pct_total | Bands | Top SOC examples |\n"
               "|---|---|---:|---:|---|---|\n")
    for naics, info in ranked:
        socs = sorted(info["socs"], key=lambda s: -s["pct"])
        examples = "; ".join(
            f"`{s['soc']}` {s['title'][:35]} ({s['pct']:.1f}%)"
            for s in socs[:3]
        )
        bands = "/".join(sorted(info["bands"]))
        out.append(f"| `{naics}` | {info['title']} | {len(info['socs'])} | "
                   f"{info['total_pct']:.1f}% | {bands} | {examples} |\n")

    # Group by NAICS-2 prefix for sector-level scanning.
    out.append("\n## Grouped by NAICS-2 prefix\n\n")
    by_prefix: dict[str, list[str]] = defaultdict(list)
    for naics in missing:
        by_prefix[naics[:2]].append(naics)
    out.append("| NAICS-2 | NAICS-2 Title | Missing NAICS-4 in this sector |\n"
               "|---|---|---:|\n")
    naics2_titles = {n: naics_titles.get(n, "") for n in by_prefix}
    for prefix in sorted(by_prefix.keys(), key=lambda p: -len(by_prefix[p])):
        title = naics2_titles.get(prefix, "")
        out.append(f"| `{prefix}` | {title} | {len(by_prefix[prefix])} |\n")

    Path(out_path).write_text("".join(out))
    print(f"Wrote {out_path} ({len(missing)} missing NAICS-4 codes)")


if __name__ == "__main__":
    main("/app/missing_naics_with_labels.md")
