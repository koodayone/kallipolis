"""
Crosswalk utilities: TOP6 → CIP → SOC chain.

Provides authoritative government mappings from California community college
programs to federal occupational classifications. Used for demand profiling
and employer-occupation alignment.

All TOP codes are 6-digit throughout. The TOP-CIP crosswalk published by the
Chancellor's Office is natively TOP6 (rows like "0101.00 - Agriculture..."
strip to "010100"); the PCAH TOP Codes to Sectors file is also TOP6. Student
calibrations are TOP6 after the student-generator refactor. No TOP4
truncation happens in this module.

Usage:
    from ontology.crosswalks import build_demand_profile
    profile = build_demand_profile("lacity", occupations, coe_demand)
"""

from __future__ import annotations

import csv
import json
import logging
from pathlib import Path

import openpyxl

logger = logging.getLogger(__name__)

# ── Data paths ────────────────────────────────────────────────────────────
#
# All five artifacts now ship in-repo. The TOP→CIP crosswalk
# (Chancellor's Office), CIP→SOC crosswalk (NCES/BLS), PCAH sector file
# (Chancellor's Office), COE demand file (California Centers of Excellence),
# and the Taxonomy of Programs vocational/CTE flag (Chancellor's Office, 7th
# Ed) are public-domain federal/state datasets on slow update
# cadences (years between revisions). Bundling them removes a dev-machine
# path dependency and lets ingestion-time code (occupations generation,
# Course→Occupation edge materialization) run inside any environment
# the backend container reaches — no cc_dataset directory required.
#
# COE_DEMAND_PATH points at the same file ontology/supply.py reads at
# runtime; previously this module referenced an out-of-repo duplicate,
# a dormant inconsistency.

_DATA_DIR = Path(__file__).parent / "data"
TOP_CIP_PATH = _DATA_DIR / "top_cip_crosswalk.csv"
CIP_SOC_PATH = _DATA_DIR / "CIP2020_SOC2018_Crosswalk.xlsx"
PCAH_SECTORS_PATH = _DATA_DIR / "TOP Codes to Sectors.xlsx"
TOP_VOCATIONAL_PATH = _DATA_DIR / "top_vocational.csv"
NAICS4_DESCRIPTIONS_PATH = _DATA_DIR / "2022_NAICS_Descriptions.xlsx"
TOP4_NAMES_PATH = _DATA_DIR / "top4_names.json"
COE_DEMAND_PATH = Path(__file__).parent / "occupational_demand_coe.csv"

CALIBRATIONS_DIR = Path(__file__).parent / "calibrations"


# ── Crosswalk loaders (cached) ────────────────────────────────────────────

_top_to_cip: dict[str, set[str]] | None = None
_cip_to_soc: dict[str, set[str]] | None = None
_coe_demand: dict[str, dict[str, dict]] | None = None
_top6_to_sector: dict[str, str] | None = None
_cte_top4_cache: set[str] | None = None
_cte_reachable_socs_cache: set[str] | None = None
_vocational_top6: set[str] | None = None
_top_titles: dict[str, str] | None = None
_naics4_titles: dict[str, str] | None = None


def _load_top_to_cip() -> dict[str, set[str]]:
    """Load TOP6 → CIP code mapping from Chancellor's Office crosswalk.

    The source CSV's TOP column has the form "0101.00 - Agriculture...";
    stripping the dot yields the 6-digit code "010100". Previously this
    was truncated to 4 digits to align with a TOP4-native student
    generator. The student generator is TOP6 now, so the full 6-digit
    resolution is preserved here.
    """
    global _top_to_cip
    if _top_to_cip is not None:
        return _top_to_cip

    mapping: dict[str, set[str]] = {}
    with open(TOP_CIP_PATH, newline="") as f:
        reader = csv.reader(f)
        next(reader)  # skip header
        for row in reader:
            top_raw = row[0].split(" - ")[0].strip().strip('"')
            cip_raw = row[2].split(" - ")[0].strip()
            top6 = top_raw.replace(".", "")
            if len(top6) == 6 and top6.isdigit() and cip_raw:
                mapping.setdefault(top6, set()).add(cip_raw)

    _top_to_cip = mapping
    logger.info(f"Loaded TOP→CIP crosswalk: {len(mapping)} TOP6 codes")
    return mapping


def load_naics4_titles() -> dict[str, str]:
    """Load NAICS-4 → industry title mapping from the Census 2022 NAICS
    Descriptions file — the canonical source for the federal industry
    classification, with one row per code at every hierarchy level.

    The OES Industry-Occupation matrix uses ad-hoc rollups (e.g.
    `3330A1` bundling 3331/3332/3334/3339) rather than a clean 4-digit
    breakout, so it covered only ~77% of the NAICS-4 codes our scraped
    employers carry. Census publishes a flat code-title table that
    covers every 4-digit industry, so we read titles from there and
    keep the OES sheet for what it does best (the HIRES_FOR set).

    Title cells carry a trailing "T" marker for trilateral NAFTA-
    harmonized industries; strip it for display.
    """
    global _naics4_titles
    if _naics4_titles is not None:
        return _naics4_titles

    titles: dict[str, str] = {}
    wb = openpyxl.load_workbook(NAICS4_DESCRIPTIONS_PATH, read_only=True)
    ws = wb.active
    if ws is None:
        _naics4_titles = titles
        return titles

    rows = ws.iter_rows(values_only=True)
    next(rows, None)  # header: Code, Title, Description
    for row in rows:
        if not row or len(row) < 2:
            continue
        code = str(row[0] or "").strip()
        title = (row[1] or "").strip() if row[1] is not None else ""
        if not (len(code) == 4 and code.isdigit() and title):
            continue
        if title.endswith("T"):
            title = title[:-1].rstrip()
        titles[code] = title

    wb.close()
    _naics4_titles = titles
    logger.info(f"Loaded NAICS-4 titles: {len(titles)} codes")
    return titles


# Cache for the TOP4/TOP2 department-name table parsed from the
# Chancellor's Office Taxonomy of Programs Manual (7th edition, May 2023).
# Populated lazily by `_load_top4_names`.
_top4_names: dict[str, dict[str, str]] | None = None


def _load_top4_names() -> dict[str, dict[str, str]]:
    """Load the TOP4 → name and TOP2 → name tables, both extracted from
    `cc-top-code-manual.pdf` and committed to `data/top4_names.json`.

    Returns `{"top4": {...}, "top2": {...}}`. The TOP4 table is the
    primary lookup; TOP2 is the fallback for TOP4 codes (like 4930)
    that don't have a standalone manual entry.
    """
    global _top4_names
    if _top4_names is not None:
        return _top4_names
    with open(TOP4_NAMES_PATH) as f:
        d = json.load(f)
    _top4_names = {"top4": d["top4"], "top2": d["top2"]}
    logger.info(
        f"Loaded TOP4 department names: {len(_top4_names['top4'])} TOP4, "
        f"{len(_top4_names['top2'])} TOP2"
    )
    return _top4_names


def top_to_department_name(top6: str | None) -> str | None:
    """Return the department display name for a course's TOP6 code,
    derived from the Chancellor's Office Taxonomy of Programs manual.

    Resolution order:
      1. The TOP4 entry in the manual (e.g., 0835 → "Physical Education").
      2. The TOP2 entry as a fallback for TOP4 codes that have no
         standalone entry (e.g., 4930 → use TOP2=49 → "Interdisciplinary
         Studies"; though `top4_names.json` overrides 4930 itself with
         the manual subheading "General Studies").
      3. None when `top6` is missing or doesn't resolve.

    This replaces the prior per-college overlay system. Every course's
    department name is sourced from a single authoritative document and
    is uniform across colleges. See
    `backend/ontology/data/top4_names.json` for the parsed table.
    """
    if not top6 or len(top6) < 4 or not top6.isdigit():
        return None
    table = _load_top4_names()
    top4 = top6[:4]
    name = table["top4"].get(top4)
    if name:
        return name
    # Fall back to TOP2 for codes without a standalone TOP4 entry.
    return table["top2"].get(top4[:2])


def load_top_titles() -> dict[str, str]:
    """Load TOP6 → title mapping from the Chancellor's Office TOP-CIP
    crosswalk file. The first column has the form
    `"0101.00 - Agriculture and Natural Resources"`; we already strip
    the dotted code in `_load_top_to_cip`, here we keep the title side.

    Used by surfaces that render TOP codes alongside their human-readable
    program-area names (the student competency profile, the partnership
    artifact's curriculum-alignment section). Falls back to bare code at
    the call site when a TOP6 has no entry.
    """
    global _top_titles
    if _top_titles is not None:
        return _top_titles

    titles: dict[str, str] = {}
    with open(TOP_CIP_PATH, newline="") as f:
        reader = csv.reader(f)
        next(reader)  # skip header
        for row in reader:
            cell = row[0].strip().strip('"')
            if " - " not in cell:
                continue
            code_part, _, title_part = cell.partition(" - ")
            top6 = code_part.strip().replace(".", "")
            title = title_part.strip()
            if len(top6) == 6 and top6.isdigit() and title:
                # First write wins; the file may carry duplicate TOP rows
                # (one per CIP it maps to) and the title is identical
                # across them.
                titles.setdefault(top6, title)

    _top_titles = titles
    logger.info(f"Loaded TOP titles: {len(titles)} TOP6 codes")
    return titles


def _load_cip_to_soc() -> dict[str, set[str]]:
    """Load CIP → SOC code mapping from NCES/BLS crosswalk."""
    global _cip_to_soc
    if _cip_to_soc is not None:
        return _cip_to_soc

    mapping: dict[str, set[str]] = {}
    wb = openpyxl.load_workbook(CIP_SOC_PATH, read_only=True)
    ws = wb["CIP-SOC"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        cip, _, soc, _ = row
        if cip and soc:
            mapping.setdefault(str(cip), set()).add(str(soc))
    wb.close()

    _cip_to_soc = mapping
    logger.info(f"Loaded CIP→SOC crosswalk: {len(mapping)} CIP codes")
    return mapping


def _load_coe_demand() -> dict[str, dict[str, dict]]:
    """Load COE occupational demand projections.

    Returns: {region: {soc_code: {annual_openings, growth_rate, median_wage, jobs}}}
    """
    global _coe_demand
    if _coe_demand is not None:
        return _coe_demand

    data: dict[str, dict[str, dict]] = {}
    with open(COE_DEMAND_PATH, newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            region = row["Region"]
            soc = row["SOC"]
            try:
                data.setdefault(region, {})[soc] = {
                    "annual_openings": int(row["Average Annual Job Openings"]),
                    "growth_rate": float(row["2024 - 2029 % Change"]),
                    "median_wage": int(row["Median Annual Earnings"]),
                    "jobs_2024": int(row["2024 Jobs"]),
                    "title": row["Description"],
                    "education": row["Typical Entry Level Education"],
                }
            except (ValueError, KeyError):
                continue

    _coe_demand = data
    logger.info(f"Loaded COE demand: {len(data)} regions, {sum(len(v) for v in data.values())} entries")
    return data


def _load_pcah_cte_top6() -> dict[str, str]:
    """Load PCAH TOP6 → sector mapping.

    The Chancellor's Office *TOP Codes to Sectors* file (PCAH) lists every
    TOP6 code classified as CTE under one of the 12 Doing-What-MATTERS /
    Strong Workforce industry sectors. This is the authoritative
    institutional definition of CTE scope for the California community
    college system.

    The PCAH file stores TOP6 codes as integers (e.g. 10100, 30220);
    formatted as six-digit strings with leading zeros ("010100",
    "030220") they align with the TOP-CIP crosswalk's key shape.

    Returns: {top6: sector_name} for every CTE-classified TOP6.
    """
    global _top6_to_sector
    if _top6_to_sector is not None:
        return _top6_to_sector

    mapping: dict[str, str] = {}
    wb = openpyxl.load_workbook(PCAH_SECTORS_PATH, read_only=True)
    ws = wb["Sheet1"]
    # Row 1 is a preamble, row 2 is the header. Data starts at row 3.
    for row in ws.iter_rows(min_row=3, values_only=True):
        top6_val, _title, sector = row
        if top6_val is None or sector is None:
            continue
        top6 = f"{int(top6_val):06d}"
        mapping[top6] = str(sector)
    wb.close()

    _top6_to_sector = mapping
    logger.info(f"Loaded PCAH CTE sectors: {len(mapping)} TOP6 codes")
    return mapping


def is_cte_top6(top6: str | None) -> bool:
    """Return True iff a TOP6 code is CTE per the PCAH TOP Codes to Sectors file.

    The PCAH file is the authoritative institutional definition of CTE
    scope for the California community college system: a TOP6 is CTE
    if and only if it appears in this file. The classification is set
    membership, not a contiguous range — there are CTE codes scattered
    across the TOP namespace (e.g. 1602 Library Technician, 0850.30
    Educational Aide) and non-CTE codes in otherwise CTE-heavy clusters.
    """
    if not top6:
        return False
    return top6 in _load_pcah_cte_top6()


def _load_vocational_top6() -> set[str]:
    """Load the set of CTE (vocational) TOP6 codes from the Chancellor's Office
    Taxonomy of Programs manual (7th Edition, May 2023).

    The manual marks each program with a leading asterisk iff it is a
    *vocational* program — the Perkins-rooted, per-TOP designation the SDCCD CTE
    Definitions Guide equates with a "CTE TOP Code". This is the authoritative,
    current, TOP6-exact definition of CTE scope.

    Sourced from `top_vocational.csv` (derived from cc-top-code-manual.pdf;
    columns: top6,title,vocational — vocational ∈ {0,1}). Returns the set of
    vocational TOP6 codes (273 of 408 in the 7th Ed).
    """
    global _vocational_top6
    if _vocational_top6 is not None:
        return _vocational_top6

    voc: set[str] = set()
    with open(TOP_VOCATIONAL_PATH, newline="") as f:
        reader = csv.reader(f)
        next(reader)  # skip header
        for row in reader:
            if len(row) >= 3 and row[2].strip() == "1":
                voc.add(row[0].strip())

    _vocational_top6 = voc
    logger.info(f"Loaded vocational (CTE) TOP6 codes: {len(voc)}")
    return voc


def is_vocational(top6: str | None) -> bool:
    """True iff a TOP6 is a CTE (vocational) program per the CCCCO Taxonomy of
    Programs manual (7th Ed) — the authoritative, TOP6-exact CTE gate.

    Preferred over the two older tests for new CTE scoping:
    - vs is_cte_top6 (PCAH sector file): ~99% the same set, but the manual is the
      primary source and is current; the PCAH file lags and is really a
      CTE→sector mapping rather than a CTE membership gate.
    - vs is_cte_top4_family: exact, so it does NOT sweep in transfer siblings
      (e.g. 083500 Physical Education, 220600 Geography, 061200 Film Studies),
      which the TOP4-family heuristic over-includes.
    """
    return bool(top6) and top6 in _load_vocational_top6()


def is_cte_top4_family(top6: str | None) -> bool:
    """True iff a TOP6's 4-digit family contains any CTE TOP6 per PCAH.

    A family-level CTE test, looser than is_cte_top6: it treats a whole TOP4
    program area as CTE when any of its TOP6 children are CTE-classified. This
    bridges TOP6-granularity gaps in the periodically-updated PCAH file — e.g.
    095690 Digital Fabrication Technician is a newer code absent from the file,
    but its 0956 Manufacturing family is unambiguously CTE, so it should count.
    Transfer/academic families with no CTE children (e.g. 0901 Engineering,
    General) are still excluded."""
    global _cte_top4_cache
    if _cte_top4_cache is None:
        _cte_top4_cache = {t[:4] for t in _load_pcah_cte_top6()}
    return bool(top6) and top6[:4] in _cte_top4_cache


def cte_reachable_socs() -> set[str]:
    """Return the set of SOC codes reachable from any PCAH-classified CTE TOP6.

    Composition: for every TOP6 classified as CTE in the PCAH file, walks
    the TOP6→CIP→SOC chain and collects the union of target SOCs. The
    result is the institutional-CTE SOC universe — an upstream input to
    the occupations generation filter.

    Cached for the process lifetime.
    """
    global _cte_reachable_socs_cache
    if _cte_reachable_socs_cache is not None:
        return _cte_reachable_socs_cache

    cte_top6 = set(_load_pcah_cte_top6().keys())
    top_cip = _load_top_to_cip()
    cip_soc = _load_cip_to_soc()

    socs: set[str] = set()
    for top6 in cte_top6 & set(top_cip.keys()):
        for cip in top_cip[top6]:
            socs.update(cip_soc.get(cip, set()))

    _cte_reachable_socs_cache = socs
    logger.info(f"Computed CTE-reachable SOC set: {len(socs)} SOCs")
    return socs


# ── CIP titles cache ──────────────────────────────────────────────────────

_cip_titles: dict[str, str] | None = None


def load_cip_titles() -> dict[str, str]:
    """Load CIP code → human-readable title from the NCES CIP-SOC crosswalk
    workbook. Used by the Curriculum Alignment pathway visualization to
    label CIP nodes with their federal taxonomy names (e.g., "15.0507 →
    Environmental/Environmental Engineering Technology/Technician").

    First-write wins: the crosswalk lists each CIP once per SOC it maps
    to, but the title is identical across those rows.
    """
    global _cip_titles
    if _cip_titles is not None:
        return _cip_titles

    titles: dict[str, str] = {}
    wb = openpyxl.load_workbook(CIP_SOC_PATH, read_only=True)
    ws = wb["CIP-SOC"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        cip_code, cip_title, _soc_code, _soc_title = row
        if cip_code and cip_title:
            # Normalize: strip dotted form to no-dot to match downstream usage.
            code_str = str(cip_code).strip()
            titles.setdefault(code_str, str(cip_title).strip().rstrip("."))
    wb.close()

    _cip_titles = titles
    logger.info(f"Loaded CIP titles: {len(titles)} CIPs")
    return titles


# ── Core functions ────────────────────────────────────────────────────────

def top6_to_soc(top6_codes: list[str]) -> dict[str, set[str]]:
    """Map TOP6 codes to SOC codes via TOP6→CIP→SOC chain.

    Returns: {top6: {soc_code, ...}} for each TOP6 that has a mapping.
    """
    top_cip = _load_top_to_cip()
    cip_soc = _load_cip_to_soc()

    result: dict[str, set[str]] = {}
    for top6 in top6_codes:
        cips = top_cip.get(top6, set())
        socs: set[str] = set()
        for cip in cips:
            socs.update(cip_soc.get(cip, set()))
        if socs:
            result[top6] = socs

    return result


def top6_to_cips(top6: str) -> list[str]:
    """Map a single TOP6 code to its CIPs via the Chancellor's Office
    TOP-CIP crosswalk. Returns a sorted list of CIP codes (empty if the
    TOP6 has no mapping). Used by the course-detail view to render the
    "Federal Curriculum Codes" section for a specific course.
    """
    top_cip = _load_top_to_cip()
    return sorted(top_cip.get(top6, set()))


def top6_to_cips_for_soc(top6: str, soc_code: str) -> list[str]:
    """Map a single TOP6 to its CIPs that also bridge to the given SOC
    via the NCES CIP-SOC crosswalk. Intersects the course's TOP6→CIP
    set with CIPs that reach the target SOC, returning a sorted list.
    Used by the course-detail view inside a partnership opportunity
    report — restricts CIPs to those consistent with the report's
    anchoring SOC, guaranteeing the course's CIP list is a subset of
    what the Curriculum Pathway hero displays.
    """
    top_cip = _load_top_to_cip()
    cip_soc = _load_cip_to_soc()
    course_cips = top_cip.get(top6, set())
    return sorted(cip for cip in course_cips if soc_code in cip_soc.get(cip, set()))


def top4_to_cips_for_soc(soc_code: str) -> dict[str, set[str]]:
    """Map each TOP4 to the CIPs that bridge it to the given SOC.

    Composes the institutional chain TOP6 → CIP → SOC down to the TOP4
    level: for each TOP4 family, the result includes the union of CIPs
    that (a) any TOP6 in that family maps to AND (b) reach the target
    SOC per the NCES CIP-SOC crosswalk. CIPs that map to the SOC but
    are unreachable from any TOP6 in the family are excluded — only
    "relevant" CIPs for this (TOP4, SOC) pair appear.

    Used by the Curriculum Alignment pathway visualization to render
    the TOP4 → CIP → SOC chain in a way that hides irrelevant CIPs
    for the SOC the report is anchored to.

    Returns: {top4: {cip_code, ...}} for every TOP4 with at least one
    bridging CIP.
    """
    top_cip = _load_top_to_cip()
    cip_soc = _load_cip_to_soc()

    # CIPs that the NCES crosswalk maps to this SOC — the destination set.
    target_cips = {cip for cip, socs in cip_soc.items() if soc_code in socs}

    # Aggregate per TOP4: union over TOP6 children, intersect with target CIPs.
    result: dict[str, set[str]] = {}
    for top6, cips in top_cip.items():
        if len(top6) < 4:
            continue
        relevant = cips & target_cips
        if not relevant:
            continue
        top4 = top6[:4]
        result.setdefault(top4, set()).update(relevant)

    return result


def build_demand_profile(
    college_key: str,
    occupations: list[dict],
    coe_region: str,
) -> list[dict]:
    """Build ranked demand profile for a college.

    Computes which occupations this college should be producing graduates
    for, ranked by composite score of enrollment weight × annual openings
    × growth × wage.

    Args:
        college_key: Pipeline key (e.g., "lacity")
        occupations: Full occupations list from occupations.json
        coe_region: COE region code (e.g., "LA", "Bay")

    Returns:
        Ranked list of dicts:
        [{soc_code, title, annual_openings, growth_rate, median_wage,
          enrollment_weight, composite_score, top6_sources}]
    """
    cal_path = CALIBRATIONS_DIR / "top6" / f"{college_key}.json"
    if not cal_path.exists():
        logger.warning(f"No TOP6 calibration for {college_key}")
        return []

    with open(cal_path) as f:
        cal = json.load(f)

    top6_data = cal.get("top6_codes", {})
    total_enrollment = cal.get("total_enrollments", 1)

    # Map TOP6 → SOC
    top6_codes = list(top6_data.keys())
    top6_soc_map = top6_to_soc(top6_codes)

    # Build SOC → enrollment weight (sum across all TOP6s that feed this SOC)
    valid_socs = {o["soc_code"] for o in occupations}
    soc_weights: dict[str, float] = {}
    soc_top6_sources: dict[str, list[str]] = {}

    for top6, socs in top6_soc_map.items():
        enrollment = top6_data.get(top6, {}).get("enrollment", 0)
        weight = enrollment / total_enrollment if total_enrollment else 0
        for soc in socs:
            if soc in valid_socs:
                soc_weights[soc] = soc_weights.get(soc, 0) + weight
                soc_top6_sources.setdefault(soc, []).append(top6)

    if not soc_weights:
        logger.warning(f"No SOC codes reachable for {college_key}")
        return []

    coe = _load_coe_demand()
    region_demand = coe.get(coe_region, {})
    if not region_demand:
        region_demand = coe.get("CA", {})
        logger.warning(f"No COE data for region {coe_region}, using statewide")

    raw_scores: list[dict] = []
    for soc, enrollment_weight in soc_weights.items():
        demand = region_demand.get(soc)
        if not demand:
            continue
        raw_scores.append({
            "soc_code": soc,
            "title": demand["title"],
            "annual_openings": demand["annual_openings"],
            "growth_rate": demand["growth_rate"],
            "median_wage": demand["median_wage"],
            "jobs_2024": demand["jobs_2024"],
            "education": demand["education"],
            "enrollment_weight": enrollment_weight,
            "top6_sources": soc_top6_sources.get(soc, []),
        })

    if not raw_scores:
        return []

    openings = [s["annual_openings"] for s in raw_scores]
    growths = [1 + s["growth_rate"] for s in raw_scores]
    wages = [s["median_wage"] for s in raw_scores]
    weights = [s["enrollment_weight"] for s in raw_scores]

    def _norm(values: list[float]) -> list[float]:
        lo, hi = min(values), max(values)
        if hi == lo:
            return [0.5] * len(values)
        return [(v - lo) / (hi - lo) for v in values]

    n_openings = _norm(openings)
    n_growths = _norm(growths)
    n_wages = _norm(wages)
    n_weights = _norm(weights)

    for i, s in enumerate(raw_scores):
        s["composite_score"] = n_weights[i] * n_openings[i] * n_growths[i] * n_wages[i]

    raw_scores.sort(key=lambda s: -s["composite_score"])
    return raw_scores
